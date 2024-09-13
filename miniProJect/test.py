import os
import unicodedata
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from matplotlib import font_manager, rc
import platform

# 한글 폰트 설정
if platform.system() == 'Darwin':  # macOS
    rc('font', family='AppleGothic')
elif platform.system() == 'Windows':  # Windows
    rc('font', family='Malgun Gothic')
else:
    print("해당 운영체제에서 한글 폰트 적용 필요")

# 경로 설정
directory = "D:/rpawork\workspace\mainpj\miniProJect/2019~2023"

# 파일 필터링 조건
def filter_files(filename):
    normalized_filename = unicodedata.normalize('NFC', filename)  # 유니코드 정규화
    return "손익계산서" in normalized_filename and "연결" not in normalized_filename and "포괄" not in normalized_filename

# 파일 리스트 필터링 및 엑셀 파일 불러오기
def get_filtered_excel_files():
    try:
        files = os.listdir(directory)
        filtered_files = [f for f in files if filter_files(f)]
        
        if filtered_files:
            print("필터링된 파일들:")
            for filename in filtered_files:
                print(filename)
            return filtered_files
        else:
            print("필터링 조건에 맞는 파일이 없습니다.")
            return []
    except FileNotFoundError:
        print(f"경로 {directory}를 찾을 수 없습니다.")
        return []

# 재무 비율 계산 함수
def calculate_financial_ratios(group):
    try:
        sales = group[group['항목코드'] == 'ifrs-full_Revenue']['당기'].values[0]
        cost_of_sales = group[group['항목코드'] == 'ifrs-full_CostOfSales']['당기'].values[0]
        sga_expenses = group[group['항목코드'] == 'dart_TotalSellingGeneralAdministrativeExpenses']['당기'].values[0]
        operating_income = group[group['항목코드'] == 'dart_OperatingIncomeLoss']['당기'].values[0]

        if sales != 0:
            cost_of_sales_ratio = (cost_of_sales / sales) * 100
            sga_expenses_ratio = (sga_expenses / sales) * 100
            operating_income_ratio = (operating_income / sales) * 100
        else:
            cost_of_sales_ratio = sga_expenses_ratio = operating_income_ratio = None

        return pd.Series({
            '회사명': group['회사명'].values[0],
            '매출원가': cost_of_sales_ratio,
            '판매비와관리비': sga_expenses_ratio,
            '영업이익': operating_income_ratio
        })
    except IndexError:
        return None

# 엑셀 파일 분석 함수
def analyze_excel_file(file_path):
    # 엑셀 파일 불러오기
    data = pd.read_excel(file_path)

    # 업종이 212인 데이터를 필터링합니다.
    filtered_data = data[data['업종'] == 212]

    # '당기' 컬럼에서 ','를 제거하고 float 타입으로 변환합니다.
    filtered_data['당기'] = filtered_data['당기'].replace(',', '', regex=True).astype(float)

    # 회사명으로 그룹화하여 재무 비율 계산을 적용합니다.
    result = filtered_data.groupby('회사명').apply(calculate_financial_ratios).dropna()

    return result

# 각 회사별로 그래프를 저장하는 함수
def save_graph(company, years, sales_list, sga_list, income_list, graph_path):
    plt.figure(figsize=(6, 4))  # 그래프 크기 설정
    plt.plot(years, sales_list, label='매출액', marker='o')
    plt.plot(years, sga_list, label='판관비', marker='o')
    plt.plot(years, income_list, label='영업이익', marker='o')
    
    plt.title(f'{company} - 연도별 재무 비율')
    plt.xlabel('연도')
    plt.ylabel('비율 (%)')
    plt.legend()
    plt.grid(True)
    plt.savefig(graph_path)
    plt.close()

# 연도별 결과를 포맷에 맞게 출력 및 엑셀에 저장하는 함수 (그래프 포함)
def save_results_to_excel_with_graph(results_by_year, output_path):
    years = sorted(results_by_year.keys())
    
    # openpyxl을 사용하여 엑셀 파일 생성
    wb = Workbook()
    ws = wb.active

    # 각 회사별로 데이터를 추출
    company_names = results_by_year[years[0]]['회사명'].unique()

    row_idx = 1  # 엑셀 행 번호

    for company in company_names:
        # 첫 번째 줄: 회사명과 연도
        ws.cell(row=row_idx, column=1, value=company)
        for col_idx, year in enumerate(years, start=2):
            ws.cell(row=row_idx, column=col_idx, value=f"{year}년")

        row_idx += 1

        # 매출원가, 판관비, 영업이익 추가
        cost_of_sales = ['매출원가']
        sga_expenses = ['판관비']
        operating_income = ['영업이익']

        sales_list = []
        sga_list = []
        income_list = []

        for year in years:
            row = results_by_year[year][results_by_year[year]['회사명'] == company]
            if not row.empty:
                cost_of_sales.append(f"{row['매출원가'].values[0]:.2f}%")
                sga_expenses.append(f"{row['판매비와관리비'].values[0]:.2f}%")
                operating_income.append(f"{row['영업이익'].values[0]:.2f}%")

                sales_list.append(row['매출원가'].values[0])
                sga_list.append(row['판매비와관리비'].values[0])
                income_list.append(row['영업이익'].values[0])
            else:
                cost_of_sales.append("N/A")
                sga_expenses.append("N/A")
                operating_income.append("N/A")
                sales_list.append(None)
                sga_list.append(None)
                income_list.append(None)

        # 데이터 엑셀에 입력
        for col_idx, val in enumerate(cost_of_sales, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        row_idx += 1

        for col_idx, val in enumerate(sga_expenses, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        row_idx += 1

        for col_idx, val in enumerate(operating_income, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        row_idx += 1

        # 그래프를 생성하고 저장
        graph_path = f"D:/rpawork\workspace\mainpj\miniProJect/{company}_graph.png"
        save_graph(company, years, sales_list, sga_list, income_list, graph_path)

        # 그래프를 엑셀에 삽입하고 크기 조절
        img = Image(graph_path)
        img.width = 150  # 너비 조정
        img.height = 100  # 높이 조정
        ws.add_image(img, f"I{row_idx - 3}")

        # 회사별로 한 줄 띄움
        row_idx += 1

    # 엑셀 파일 저장
    wb.save(output_path)
    print(f"결과가 엑셀 파일로 저장되었습니다: {output_path}")

# 메인 함수 실행
if __name__ == "__main__":
    filtered_files = get_filtered_excel_files()  # 필터링된 엑셀 파일 목록 가져오기
    results_by_year = {}

    for file in filtered_files:
        # 파일 이름에서 연도를 추출 (파일 이름의 첫 번째 부분이 연도라고 가정)
        year = file.split('_')[0]
        
        # 엑셀 파일 경로 설정
        file_path = os.path.join(directory, file)

        # 엑셀 파일을 분석하고 연도별로 결과를 저장
        result = analyze_excel_file(file_path)
        results_by_year[year] = result

    # 엑셀 파일 경로 입력
    output_path = "D:/rpawork\workspace\mainpj\miniProJect/result.xlsx"

    # 연도별 분석 결과를 엑셀에 저장 (그래프 포함)
    save_results_to_excel_with_graph(results_by_year, output_path)
